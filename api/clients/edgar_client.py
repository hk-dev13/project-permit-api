from __future__ import annotations

import os
import logging
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook  # type: ignore

logger = logging.getLogger(__name__)


class EDGARClient:
    """Loader for EDGAR UCDB emissions Excel (EDGAR_emiss_on_UCDB_2024.xlsx).

    Provides aggregation to country-year totals per pollutant by summing all
    Urban Centre (UC) rows for the same UC_country across sectors.

    Notes:
      - Pollutants available include CO2, GWP_100_AR5_GHG, PM2.5, NOx.
      - Units: ton/year (per workbook metadata).
      - Granularity is urban; totals reflect urban emissions only.
    """

    # Global cache shared across instances: key -> {agg_by_country, header, colmap, country_col_idx}
    _GLOBAL_CACHE: Dict[str, Dict[str, Any]] = {}

    def __init__(self, xlsx_path: Optional[str] = None) -> None:
        default_path = os.path.join(os.getcwd(), "reference", "EDGAR_emiss_on_UCDB_2024.xlsx")
        self.xlsx_path = (xlsx_path or os.getenv("EDGAR_XLSX_PATH") or default_path)
        self.sheet_name_candidates = [
            "EDGAR_emiss_on_UCDB_2024",
            "EDGA R_emiss_on_UCDB_2024".replace(" ", ""),  # safety
        ]
        # Caches
        self._header: Optional[List[str]] = None
        self._colmap: Optional[Dict[Tuple[str, int], List[int]]] = None
        self._country_col_idx: Optional[int] = None
        self._agg_by_country: Optional[Dict[str, Dict[str, Dict[int, float]]]] = None

    # ---- Internal helpers ----
    def _cache_key(self) -> str:
        try:
            mtime = os.path.getmtime(self.xlsx_path)
        except Exception:
            mtime = "na"
        return f"{self.xlsx_path}:{mtime}"

    def _load_sheet(self):
        if not os.path.exists(self.xlsx_path):
            raise FileNotFoundError(f"EDGAR file not found: {self.xlsx_path}")
        wb = load_workbook(self.xlsx_path, read_only=True, data_only=True)
        # pick main data sheet
        ws = None
        for name in wb.sheetnames:
            if name.strip() in ("EDGAR_emiss_on_UCDB_2024",):
                ws = wb[name]
                break
        if ws is None:
            # fallback to last sheet which often is the data sheet
            ws = wb[wb.sheetnames[-1]]
        return wb, ws

    def _parse_header(self, header_row: List[Any]) -> None:
        header = [str(c).strip() if c is not None else "" for c in header_row]
        self._header = header
        # Locate UC_country column
        try:
            self._country_col_idx = header.index("UC_country")
        except ValueError:
            # try case-insensitive search
            self._country_col_idx = next((i for i, h in enumerate(header) if h.strip().lower() == "uc_country"), 2)

        # Build (pollutant, year) -> [col_indices] map (sum across sectors)
        colmap: Dict[Tuple[str, int], List[int]] = {}
        for idx, h in enumerate(header):
            if not h or not h.startswith("EMI_"):
                continue
            parts = h.split("_")
            if len(parts) < 4:
                continue
            # Expect structure: EMI_<pollutant...>_<sector>_<year>
            year_token = parts[-1]
            sector = parts[-2]  # unused but validated by structure
            pollutant = "_".join(parts[1:-2])
            try:
                year = int(year_token)
            except Exception:
                continue
            key = (pollutant, year)
            colmap.setdefault(key, []).append(idx)

        self._colmap = colmap

    def _ensure_aggregated(self) -> None:
        if self._agg_by_country is not None:
            return
        # Use global cache if available
        key = self._cache_key()
        cached = self._GLOBAL_CACHE.get(key)
        if cached is not None:
            self._agg_by_country = cached.get("agg_by_country")
            self._header = cached.get("header")
            self._colmap = cached.get("colmap")
            self._country_col_idx = cached.get("country_col_idx")
            return

        wb, ws = self._load_sheet()
        try:
            rows = ws.iter_rows(min_row=1, values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                raise ValueError("EDGAR sheet is empty")
            self._parse_header(list(header))
            if self._colmap is None or self._country_col_idx is None:
                raise ValueError("Failed to parse EDGAR header")

            agg: Dict[str, Dict[str, Dict[int, float]]] = {}
            for row in rows:
                if row is None:
                    continue
                try:
                    country = str(row[self._country_col_idx] or "").strip()
                except Exception:
                    country = ""
                if not country:
                    continue

                # ensure country bucket
                bucket = agg.setdefault(country, {})
                # Sum per (pollutant, year) across sectors
                for (pollutant, year), col_idxs in self._colmap.items():
                    total = 0.0
                    any_val = False
                    for ci in col_idxs:
                        try:
                            v = row[ci]
                            if isinstance(v, (int, float)):
                                total += float(v)
                                any_val = True
                            elif isinstance(v, str) and v.strip() != "":
                                total += float(v)
                                any_val = True
                        except Exception:
                            continue
                    if any_val:
                        polmap = bucket.setdefault(pollutant, {})
                        polmap[year] = polmap.get(year, 0.0) + total

            self._agg_by_country = agg
            # Save to global cache
            self._GLOBAL_CACHE[key] = {
                "agg_by_country": self._agg_by_country,
                "header": self._header,
                "colmap": self._colmap,
                "country_col_idx": self._country_col_idx,
            }
        finally:
            try:
                wb.close()
            except Exception:
                pass

    # ---- Public API ----
    def get_country_series(self, country: str, pollutant: str) -> List[Dict[str, Any]]:
        """Return sorted series for a country and pollutant: [{year, value}]."""
        if not country:
            return []
        self._ensure_aggregated()
        series: List[Dict[str, Any]] = []
        agg = self._agg_by_country or {}
        data = (agg.get(country) or {}).get(pollutant) or {}
        for y, v in data.items():
            series.append({"year": int(y), "value": float(v)})
        series.sort(key=lambda r: r["year"])  # ascending
        return series

    def compute_country_trend(self, country: str, pollutant: str = "PM2.5", window: int = 3) -> Dict[str, Any]:
        """Compute simple trend over the last `window` points for a pollutant.

        Returns {"pollutant": pollutant, "slope": float, "increase": bool, "years": [..]}
        """
        series = self.get_country_series(country, pollutant)
        if len(series) < 2:
            return {"pollutant": pollutant, "slope": 0.0, "increase": False, "years": []}
        sel = series[-window:] if len(series) >= window else series
        slope = float(sel[-1]["value"] - sel[0]["value"])  # simple delta
        return {
            "pollutant": pollutant,
            "slope": slope,
            "increase": slope > 0.0,
            "years": [r["year"] for r in sel],
        }

    # Backward/explicit helper name requested in requirements
    def get_country_emissions_trend(self, country: str, pollutant: str = "PM2.5", window: int = 3) -> Dict[str, Any]:
        """Alias helper returning trend dict for country and pollutant.

        Example output: {"pollutant": "PM2.5", "slope": 12.3, "increase": True, "years": [2015, 2020, 2022]}
        """
        return self.compute_country_trend(country, pollutant=pollutant, window=window)


__all__ = ["EDGARClient"]
