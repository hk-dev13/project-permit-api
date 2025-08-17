from __future__ import annotations

import os
import logging
from typing import Any, Dict, List, Optional
import csv
import io

import requests
from openpyxl import load_workbook  # type: ignore

from api.utils.schema import ensure_iso_cert_schema

logger = logging.getLogger(__name__)


class ISOClient:
    """Client for ISO 14001 certifications (scaffold with sample fallback).

    If ISO_API_BASE is unset, returns sample data.
    """

    def __init__(self) -> None:
        self.api_base = os.getenv("ISO_API_BASE", "").rstrip("/")
        self.csv_url = os.getenv("ISO_CSV_URL", "").strip()
        self.xlsx_path = os.getenv("ISO_XLSX_PATH", os.path.join(os.getcwd(), "reference", "list_iso.xlsx"))
        self.session = requests.Session()
        self.session.headers.update({
            "Accept": "application/json",
            "User-Agent": "project-permit-api/1.0 (+https://github.com/hk-dev13)"
        })

    def create_sample_data(self) -> List[Dict[str, Any]]:
        return [
            {"company": "Green Energy Co", "country": "US", "certificate": "ISO 14001", "valid_until": "2026-12-31"},
            {"company": "Eco Manufacturing GmbH", "country": "DE", "certificate": "ISO 14001", "valid_until": "2025-08-01"},
            {"company": "Sustain PT", "country": "ID", "certificate": "ISO 14001", "valid_until": "2027-01-15"},
        ]

    def _load_from_csv_or_json(self, url: str) -> List[Dict[str, Any]]:
        try:
            resp = self.session.get(url, timeout=30)
            ct = (resp.headers.get("Content-Type") or "").lower()
            text = resp.text
            if "json" in ct or (text.lstrip().startswith("[") or text.lstrip().startswith("{")):
                data = resp.json()
                if isinstance(data, dict):
                    # common wrapper key
                    data = data.get("data", [])
                if not isinstance(data, list):
                    raise ValueError("Unexpected JSON shape for ISO dataset")
                return [d for d in data if isinstance(d, dict)]
            # assume CSV
            buf = io.StringIO(text)
            reader = csv.DictReader(buf)
            return [dict(row) for row in reader]
        except Exception as e:
            logger.error(f"ISO CSV/JSON load error: {e}")
            return []

    def _load_from_excel(self, path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """Load ISO 14001 list from Excel. Scans for a sheet and header row containing 'Company'."""
        rows: List[Dict[str, Any]] = []
        try:
            if not os.path.exists(path):
                return rows
            wb = load_workbook(path, read_only=True, data_only=True)
            sh = sheet_name
            if not sh:
                # Prefer a sheet that looks like a list of certified companies
                for name in wb.sheetnames:
                    if "iso 14001" in name.lower() and "certified" in name.lower():
                        sh = name
                        break
                if not sh:
                    sh = wb.sheetnames[0]
            ws = wb[sh]
            # find header row: the first row containing 'Company'
            header = None
            header_row_index = 0
            for idx, r in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
                vals = ["" if c is None else str(c).strip() for c in r]
                if any(v for v in vals) and any("company" == v.lower() for v in vals):
                    header = vals
                    header_row_index = idx
                    break
            if not header:
                return rows
            # Build column map
            def col_idx(name: str) -> Optional[int]:
                lname = name.lower()
                for i, h in enumerate(header):
                    if h.lower() == lname:
                        return i
                return None
            idx_company = col_idx("Company")
            idx_eff = col_idx("Effective date")
            idx_exp = col_idx("Expiry date")
            # iterate following rows
            for r in ws.iter_rows(min_row=header_row_index + 1, values_only=True):
                cells = [None if c is None else c for c in r]
                # stop if empty row
                if not any(c is not None and str(c).strip() for c in cells):
                    continue
                comp = str(cells[idx_company]).strip() if (idx_company is not None and idx_company < len(cells) and cells[idx_company] is not None) else ""
                if not comp or comp.lower().startswith("appendix"):
                    # skip title rows
                    continue
                def dt_to_str(val: Any) -> Optional[str]:
                    try:
                        from datetime import datetime
                        if hasattr(val, "strftime"):
                            return val.strftime("%Y-%m-%d")
                        s = str(val).strip()
                        if not s:
                            return None
                        # try parse y-m-d or y/m/d
                        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
                            try:
                                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
                            except Exception:
                                pass
                        return s
                    except Exception:
                        return None
                valid_until = None
                if idx_exp is not None and idx_exp < len(cells):
                    valid_until = dt_to_str(cells[idx_exp])
                record = {
                    "company": comp,
                    "country": None,  # not available in this sheet
                    "certificate": "ISO 14001",
                    "valid_until": valid_until,
                    "_source_sheet": sh,
                }
                rows.append(record)
        except Exception as e:
            logger.error(f"ISO Excel load error: {e}")
        return rows

    def get_iso14001_certifications(self, *, country: Optional[str] = None, limit: int = 100) -> List[Dict[str, Any]]:
        data: List[Dict[str, Any]] = []
        # Prefer explicit CSV URL if provided
        if self.csv_url:
            data = self._load_from_csv_or_json(self.csv_url)
        # Also merge Excel list if available
        excel_rows = self._load_from_excel(self.xlsx_path)
        if excel_rows:
            data.extend(excel_rows)
        elif self.api_base and not data:
            # Placeholder for future real API call
            try:
                url = f"{self.api_base}/iso1401"  # adjust when real endpoint available
                params = {"country": country} if country else {}
                resp = self.session.get(url, params=params, timeout=30)
                if resp.status_code == 200:
                    data = resp.json()
                    if not isinstance(data, list):
                        raise ValueError("Unexpected ISO response shape")
                else:
                    logger.warning(f"ISO API HTTP {resp.status_code}, using sample data")
                    data = self.create_sample_data()
            except Exception as e:
                logger.error(f"ISO API error: {e}")
                data = self.create_sample_data()
        if not data:
            data = self.create_sample_data()

        if country:
            data = [d for d in data if str(d.get("country", "")).upper() == country.upper()]
        if limit and len(data) > limit:
            data = data[:limit]
        return [ensure_iso_cert_schema(rec) for rec in data if isinstance(rec, dict)]


__all__ = ["ISOClient"]
