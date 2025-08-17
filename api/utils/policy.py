from __future__ import annotations

import os
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook  # type: ignore


DEFAULT_POLICY_XLSX = os.path.join(
    os.getcwd(), "reference", "Annex III_Best practices and justifications.xlsx"
)


def _normalize_header(vals: List[Any]) -> List[str]:
    names = []
    for v in vals:
        s = "" if v is None else str(v).strip()
        names.append(s)
    return names


def load_best_practices(path: Optional[str] = None, sheet_name: str = "Best practices") -> List[Dict[str, Any]]:
    """Load best practices rows from the Excel file.

    Returns a list of dicts with normalized keys: id, country, typology, legislative_reference,
    level, scheme, description, scope, justification, valid_emas_feature, extra_info.
    """
    p = path or os.getenv("POLICY_XLSX_PATH") or DEFAULT_POLICY_XLSX
    if not os.path.exists(p):
        return []
    wb = load_workbook(p, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        # fallback to first sheet
        ws = wb[wb.sheetnames[0]]
    else:
        ws = wb[sheet_name]
    # Header is row 1
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = _normalize_header(list(header_row))
    # Map expected columns
    def idx(name: str) -> Optional[int]:
        lname = name.lower()
        for i, h in enumerate(headers):
            if h.lower() == lname:
                return i
        return None
    col_map = {
        "id": idx("ID"),
        "country": idx("Country"),
        "typology": idx("Typology"),
        "legislative_reference": idx("Legislative Reference"),
        "level": idx("Level of application"),
        "scheme": idx("Voluntary scheme addressed"),
        "description": idx("Description"),
        "scope": idx("Scope"),
        "justification": idx("Justification"),
        "valid_emas_feature": idx("Valid based on an EMAS feature? "),
        "extra_info": idx("Extra info required"),
    }
    rows: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        vals = list(r)
        if not any(v is not None and str(v).strip() for v in vals):
            continue
        def get_col(k: str) -> Optional[str]:
            i = col_map.get(k)
            if i is None or i >= len(vals):
                return None
            v = vals[i]
            return None if v is None else str(v).strip()
        rec = {k: get_col(k) for k in col_map.keys()}
        rows.append(rec)
    return rows


def practices_for_country(practices: List[Dict[str, Any]], country: str) -> List[Dict[str, Any]]:
    cl = (country or "").strip().lower()
    return [p for p in practices if (p.get("country") or "").strip().lower() == cl]
